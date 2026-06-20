"""
PaymentSummary Organizer
- แยก Sheet ตาม Customer No.
- แยก Sheet สำหรับ Payment Method: เช็คล่วงหน้า และ เช็คล่วงหน้า-onhand
- เพิ่ม SUM ยอดรวมในแต่ละ Sheet
"""

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import glob

# ========== หา input file ==========
script_dir = os.path.dirname(os.path.abspath(__file__))

# ลอง path โดยตรงจากไฟล์ที่อัปโหลดใน Claude
CLAUDE_UPLOAD_PATHS = [
    r"C:\Users\user\AppData\Roaming\Claude\local-agent-mode-sessions\6b664f92-cd4e-43a8-8c27-6b1ec65dc453\e21b356c-ae08-4854-91d4-ec23d5dc1dde\local_65a01780-52c9-4fcb-90c7-5e875505c73d\uploads\PaymentSummary (13)-b1bf7457.xls",
    r"C:\Users\user\AppData\Roaming\Claude\local-agent-mode-sessions\6b664f92-cd4e-43a8-8c27-6b1ec65dc453\e21b356c-ae08-4854-91d4-ec23d5dc1dde\local_65a01780-52c9-4fcb-90c7-5e875505c73d\uploads\PaymentSummary (13).xls",
]

xls_files = []
for p in CLAUDE_UPLOAD_PATHS:
    if os.path.exists(p):
        xls_files = [p]
        break

if not xls_files:
    xls_files = glob.glob(os.path.join(script_dir, "PaymentSummary*.xls")) + \
                glob.glob(os.path.join(script_dir, "PaymentSummary*.xlsx"))

if not xls_files:
    import pathlib
    home = pathlib.Path.home()
    found = list(home.rglob("uploads/PaymentSummary*.xls*"))
    if found:
        xls_files = [str(found[0])]

if not xls_files:
    print("❌ ไม่พบไฟล์ PaymentSummary*.xls หรือ .xlsx")
    print("   กรุณาวางไฟล์ไว้ในโฟลเดอร์เดียวกับ script นี้แล้วรันใหม่")
    input("กด Enter เพื่อออก...")
    exit()

input_file = xls_files[0]
print(f"📂 อ่านไฟล์: {input_file}")

# ========== อ่านข้อมูล ==========
all_sheets = pd.read_excel(input_file, sheet_name=None, header=None)
print(f"   พบ {len(all_sheets)} sheet(s): {list(all_sheets.keys())}")

# เลือก sheet แรก (หรือ sheet ที่มีข้อมูล)
main_sheet_name = list(all_sheets.keys())[0]
raw = all_sheets[main_sheet_name]

print(f"\n📋 ข้อมูลดิบ ({raw.shape[0]} แถว x {raw.shape[1]} คอลัมน์):")
print(raw.head(10).to_string())

# ========== หา header row ==========
header_row = None
for i, row in raw.iterrows():
    row_str = " ".join([str(v).strip() for v in row if pd.notna(v)]).lower()
    if any(kw in row_str for kw in ["customer", "payment", "cheque", "amount", "เช็ค", "ลูกหนี้", "จำนวน"]):
        header_row = i
        print(f"\n✅ พบ header ที่แถว {i}: {list(row)}")
        break

if header_row is None:
    # ลองแถวแรกที่มีข้อมูลหลายคอลัมน์
    for i, row in raw.iterrows():
        non_null = row.dropna()
        if len(non_null) >= 3:
            header_row = i
            print(f"\n⚠️  ใช้แถว {i} เป็น header (ไม่พบ keyword ที่ชัดเจน): {list(row)}")
            break

if header_row is None:
    header_row = 0
    print(f"\n⚠️  ใช้แถว 0 เป็น header (default)")

# อ่านใหม่โดยใช้ header ที่ถูกต้อง
df = pd.read_excel(input_file, sheet_name=main_sheet_name, header=header_row)
df.columns = [str(c).strip() for c in df.columns]

# ลบแถวที่ว่างทั้งหมด
df.dropna(how='all', inplace=True)
df.reset_index(drop=True, inplace=True)

print(f"\n📊 คอลัมน์ที่พบ: {list(df.columns)}")
print(f"   จำนวนแถวข้อมูล: {len(df)}")
print(df.head(5).to_string())

# ========== ระบุคอลัมน์ ==========
cols = df.columns.tolist()

def find_col(keywords):
    for kw in keywords:
        for c in cols:
            if kw.lower() in c.lower():
                return c
    return None

cust_col    = find_col(["customer no", "cust no", "custno", "customer_no", "ลูกหนี้", "รหัสลูกค้า"])
method_col  = find_col(["payment method", "paymethod", "pay method", "วิธีชำระ", "ประเภทเช็ค", "method"])
amount_col  = find_col(["amount", "จำนวนเงิน", "ยอด", "total"])
name_col    = find_col(["name", "ชื่อ", "customer name"])

print(f"\n🔍 คอลัมน์ที่จับคู่ได้:")
print(f"   Customer No : {cust_col}")
print(f"   Payment Method: {method_col}")
print(f"   Amount      : {amount_col}")
print(f"   Name        : {name_col}")

# ถ้าหาไม่ได้ ใช้คอลัมน์แรกๆ
if cust_col is None and len(cols) > 0:
    cust_col = cols[0]
    print(f"   ⚠️  ใช้คอลัมน์แรก '{cust_col}' เป็น Customer No")
if method_col is None and len(cols) > 1:
    method_col = cols[1]
    print(f"   ⚠️  ใช้คอลัมน์ที่ 2 '{method_col}' เป็น Payment Method")

# ========== สไตล์ ==========
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUB_FILL      = PatternFill("solid", fgColor="2E75B6")
SUB_FONT      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TOTAL_FILL    = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FONT    = Font(name="Arial", bold=True, color="1F4E79", size=10)
DATA_FONT     = Font(name="Arial", size=10)
ALT_FILL      = PatternFill("solid", fgColor="EBF5FF")
NORMAL_FILL   = PatternFill("solid", fgColor="FFFFFF")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT         = Alignment(horizontal="right",  vertical="center")
THIN          = Side(border_style="thin", color="BDD7EE")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill = fill; cell.font = font
    cell.alignment = CENTER; cell.border = BORDER

def style_data(cell, row_idx, is_num=False):
    cell.font = DATA_FONT
    cell.fill = ALT_FILL if row_idx % 2 == 0 else NORMAL_FILL
    cell.alignment = RIGHT if is_num else LEFT
    cell.border = BORDER

def style_total(cell):
    cell.fill = TOTAL_FILL; cell.font = TOTAL_FONT
    cell.alignment = RIGHT; cell.border = BORDER

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val.encode('utf-8')))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len * 0.7 + 2, 10), 40)

def write_df_to_sheet(ws, data_df, title, sheet_type=""):
    """เขียน DataFrame ลง worksheet พร้อมจัดรูปแบบ"""
    if data_df.empty:
        ws.cell(1, 1, f"ไม่มีข้อมูลสำหรับ {title}")
        return

    ncols = len(data_df.columns)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(1, 1, title)
    tc.fill = HEADER_FILL
    tc.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 28

    if sheet_type:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        sc = ws.cell(2, 1, sheet_type)
        sc.fill = SUB_FILL; sc.font = SUB_FONT
        sc.alignment = CENTER
        ws.row_dimensions[2].height = 20
        hdr_row = 3
    else:
        hdr_row = 2

    # Header
    ws.row_dimensions[hdr_row].height = 22
    for ci, col_name in enumerate(data_df.columns, 1):
        style_header(ws.cell(hdr_row, ci, col_name))

    # Data rows
    num_cols = set()
    for ci, col_name in enumerate(data_df.columns, 1):
        try:
            if data_df[col_name].dtype in ['float64','int64']:
                num_cols.add(ci)
        except:
            pass

    data_start = hdr_row + 1
    for ri, (_, row) in enumerate(data_df.iterrows()):
        ws.row_dimensions[data_start + ri].height = 18
        for ci, val in enumerate(row, 1):
            c = ws.cell(data_start + ri, ci)
            if pd.isna(val):
                c.value = None
            elif ci in num_cols:
                try:
                    c.value = float(val)
                    c.number_format = '#,##0.00'
                except:
                    c.value = val
            else:
                c.value = val
            style_data(c, ri, ci in num_cols)

    # SUM row
    total_row = data_start + len(data_df)
    ws.row_dimensions[total_row].height = 22
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=max(1, len(data_df.columns)-len(num_cols)))
    tc2 = ws.cell(total_row, 1, "รวมทั้งหมด")
    style_total(tc2)

    for ci in num_cols:
        col_letter = get_column_letter(ci)
        sum_cell = ws.cell(total_row, ci)
        sum_cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{total_row-1})"
        sum_cell.number_format = '#,##0.00'
        style_total(sum_cell)

    # Freeze panes
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    auto_width(ws)

# ========== สร้าง Workbook ==========
wb = openpyxl.Workbook()
wb.remove(wb.active)  # ลบ sheet ว่าง

# --- 1. Sheet รวม (All Data) ---
ws_all = wb.create_sheet("ข้อมูลทั้งหมด")
write_df_to_sheet(ws_all, df, "รายงานสรุปการชำระเงิน - ข้อมูลทั้งหมด")
print("\n✅ สร้าง Sheet: ข้อมูลทั้งหมด")

# --- 2. Sheet แยกตาม Payment Method ---
pm_keywords = {
    "เช็คล่วงหน้า-onhand": ["onhand", "on-hand", "on hand", "เช็คล่วงหน้า-onhand"],
    "เช็คล่วงหน้า":         ["เช็คล่วงหน้า"],
}

created_pm_sheets = []
if method_col:
    for pm_name, pm_keys in pm_keywords.items():
        mask = pd.Series([False] * len(df))
        for kw in pm_keys:
            mask |= df[method_col].astype(str).str.lower().str.contains(kw.lower(), na=False)
        pm_df = df[mask].copy()

        safe_name = pm_name[:31]
        ws_pm = wb.create_sheet(safe_name)
        write_df_to_sheet(ws_pm, pm_df,
                          f"รายงานสรุปการชำระเงิน",
                          f"Payment Method: {pm_name} ({len(pm_df)} รายการ)")
        created_pm_sheets.append((pm_name, len(pm_df)))
        print(f"✅ สร้าง Sheet: {safe_name} ({len(pm_df)} รายการ)")
else:
    print("⚠️  ไม่พบคอลัมน์ Payment Method ข้ามส่วนนี้")

# --- 3. Sheets แยกตาม Customer No. ---
if cust_col:
    cust_list = df[cust_col].dropna().unique()
    cust_list = sorted([str(c).strip() for c in cust_list if str(c).strip() not in ["", "nan"]])
    print(f"\n📦 พบ Customer No. ทั้งหมด {len(cust_list)} ราย: {cust_list[:10]}{'...' if len(cust_list)>10 else ''}")

    for cust in cust_list:
        cust_df = df[df[cust_col].astype(str).str.strip() == cust].copy()
        # ชื่อ sheet: ตัด 31 ตัวอักษร
        sheet_name = f"C-{cust}"[:31]
        # ถ้าชื่อซ้ำ
        base = sheet_name
        idx = 1
        while sheet_name in [ws.title for ws in wb.worksheets]:
            sheet_name = f"{base[:28]}_{idx}"
            idx += 1

        cust_name = ""
        if name_col and name_col in cust_df.columns:
            names = cust_df[name_col].dropna().unique()
            if len(names) > 0:
                cust_name = str(names[0])

        ws_cust = wb.create_sheet(sheet_name)
        title_str = f"ลูกหนี้: {cust}"
        if cust_name:
            title_str += f" - {cust_name}"
        write_df_to_sheet(ws_cust, cust_df, title_str,
                          f"Customer No: {cust}  |  {len(cust_df)} รายการ")
        print(f"   ✅ Sheet: {sheet_name} ({len(cust_df)} รายการ)")
else:
    print("⚠️  ไม่พบคอลัมน์ Customer No. ข้ามส่วนนี้")

# --- 4. Sheet สรุปภาพรวม (Summary) ---
ws_sum = wb.create_sheet("สรุปภาพรวม", 0)  # วางไว้หน้าสุด
ws_sum.sheet_view.showGridLines = False

# Title
ws_sum.merge_cells("A1:F1")
t = ws_sum.cell(1, 1, "📊 สรุปภาพรวม PaymentSummary")
t.fill = HEADER_FILL; t.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
t.alignment = CENTER
ws_sum.row_dimensions[1].height = 32

# Sub header
ws_sum.merge_cells("A2:F2")
from datetime import datetime
s = ws_sum.cell(2, 1, f"สร้างเมื่อ: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
s.fill = SUB_FILL; s.font = SUB_FONT; s.alignment = CENTER

# Section: Payment Method
row = 4
ws_sum.cell(row, 1, "Payment Method").fill = HEADER_FILL
ws_sum.cell(row, 1).font = HEADER_FONT
ws_sum.cell(row, 1).alignment = CENTER
ws_sum.cell(row, 2, "จำนวนรายการ").fill = HEADER_FILL
ws_sum.cell(row, 2).font = HEADER_FONT
ws_sum.cell(row, 2).alignment = CENTER
if amount_col:
    ws_sum.cell(row, 3, "ยอดรวม").fill = HEADER_FILL
    ws_sum.cell(row, 3).font = HEADER_FONT
    ws_sum.cell(row, 3).alignment = CENTER

row += 1
if method_col:
    for ri, (pm_name, pm_keys) in enumerate(pm_keywords.items()):
        mask = pd.Series([False] * len(df))
        for kw in pm_keys:
            mask |= df[method_col].astype(str).str.lower().str.contains(kw.lower(), na=False)
        pm_df = df[mask]
        c1 = ws_sum.cell(row, 1, pm_name)
        c2 = ws_sum.cell(row, 2, len(pm_df))
        c1.font = DATA_FONT; c2.font = DATA_FONT
        c1.fill = ALT_FILL if ri % 2 == 0 else NORMAL_FILL
        c2.fill = ALT_FILL if ri % 2 == 0 else NORMAL_FILL
        c1.border = BORDER; c2.border = BORDER
        c1.alignment = LEFT; c2.alignment = RIGHT
        if amount_col and amount_col in pm_df.columns:
            total = pd.to_numeric(pm_df[amount_col], errors='coerce').sum()
            c3 = ws_sum.cell(row, 3, total)
            c3.number_format = '#,##0.00'
            c3.font = DATA_FONT
            c3.fill = ALT_FILL if ri % 2 == 0 else NORMAL_FILL
            c3.border = BORDER; c3.alignment = RIGHT
        row += 1

# Section: Customer No.
row += 1
ws_sum.cell(row, 1, "Customer No.").fill = HEADER_FILL
ws_sum.cell(row, 1).font = HEADER_FONT; ws_sum.cell(row, 1).alignment = CENTER
ws_sum.cell(row, 2, "ชื่อลูกค้า").fill = HEADER_FILL
ws_sum.cell(row, 2).font = HEADER_FONT; ws_sum.cell(row, 2).alignment = CENTER
ws_sum.cell(row, 3, "จำนวนรายการ").fill = HEADER_FILL
ws_sum.cell(row, 3).font = HEADER_FONT; ws_sum.cell(row, 3).alignment = CENTER
if amount_col:
    ws_sum.cell(row, 4, "ยอดรวม").fill = HEADER_FILL
    ws_sum.cell(row, 4).font = HEADER_FONT; ws_sum.cell(row, 4).alignment = CENTER
row += 1

if cust_col:
    for ri, cust in enumerate(cust_list):
        cust_df = df[df[cust_col].astype(str).str.strip() == cust]
        c1 = ws_sum.cell(row, 1, cust)
        c3 = ws_sum.cell(row, 3, len(cust_df))
        fill = ALT_FILL if ri % 2 == 0 else NORMAL_FILL
        for c in [c1, c3]:
            c.font = DATA_FONT; c.fill = fill; c.border = BORDER
        c1.alignment = CENTER; c3.alignment = RIGHT

        cust_name_val = ""
        if name_col and name_col in cust_df.columns:
            names = cust_df[name_col].dropna().unique()
            if len(names) > 0:
                cust_name_val = str(names[0])
        c2 = ws_sum.cell(row, 2, cust_name_val)
        c2.font = DATA_FONT; c2.fill = fill; c2.border = BORDER; c2.alignment = LEFT

        if amount_col and amount_col in cust_df.columns:
            total = pd.to_numeric(cust_df[amount_col], errors='coerce').sum()
            c4 = ws_sum.cell(row, 4, total)
            c4.number_format = '#,##0.00'
            c4.font = DATA_FONT; c4.fill = fill; c4.border = BORDER; c4.alignment = RIGHT
        row += 1

for col_idx in range(1, 6):
    ws_sum.column_dimensions[get_column_letter(col_idx)].width = 22
ws_sum.row_dimensions[1].height = 32
ws_sum.freeze_panes = "A4"

# ========== บันทึกไฟล์ ==========
output_path = os.path.join(script_dir, "PaymentSummary_Organized.xlsx")
wb.save(output_path)
print(f"\n🎉 บันทึกไฟล์เสร็จแล้ว: {output_path}")
print(f"   📄 Sheets ที่สร้าง: {[ws.title for ws in wb.worksheets]}")
input("\nกด Enter เพื่อออก...")
